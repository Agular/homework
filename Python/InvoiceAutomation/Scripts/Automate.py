import openpyxl as pyxl

# Load the xlsx files ("workbooks")
proforma = pyxl.load_workbook("proforma.xlsx")
source = pyxl.load_workbook("here1.xlsx")
# Get the active sheets.
ws = proforma.active
ws1 = source.active
for i in range(2, 122):
    # B Column is for Entity
    ws['H5'].value = ws1['B' + str(i)].value
    # D is for address
    ws['H6'].value = ws1['D' + str(i)].value
    # E is for Suite (if given, otherwise blank)
    ws['H7'].value = ws1['E' + str(i)].value
    # G is for postal code and F is for City
    ws['H8'].value = ws1['G' + str(i)].value + " " + ws1['F' + str(i)].value
    # H Column is for Country
    ws['H9'].value = ws1['H' + str(i)].value
    # I Is for contact name
    ws['H10'].value = "Attn: " + ws1['I' + str(i)].value
    # J is for Telephone number
    ws['H11'].value = ws1['J' + str(i)].value
    # K is for email.
    ws['H12'].value = ws1['K' + str(i)].value
    # L is for quantity
    ws['F24'].value = ws1['L' + str(i)].value
    ws['F25'].value = ws1['L' + str(i)].value
    ws['F26'].value = ws1['L' + str(i)].value
    proforma.save(ws1['F'+str(i)].value + ".xlsx")
